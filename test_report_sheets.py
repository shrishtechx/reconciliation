"""Verify the Excel report has no Audit Log sheet and data matches API results."""
import requests
import openpyxl
import io
import json

base = "http://localhost:8000"

# Load sample and reconcile
requests.post(f"{base}/api/sample")
requests.post(f"{base}/api/reconcile")

# Get API results
api_results = requests.get(f"{base}/api/results").json()

# Download report
report = requests.get(f"{base}/api/report")
wb = openpyxl.load_workbook(io.BytesIO(report.content))

print(f"Sheets in report: {wb.sheetnames}")
assert "Audit Log" not in wb.sheetnames, "Audit Log sheet should NOT exist!"
assert "Summary" in wb.sheetnames
assert "Matched Transactions" in wb.sheetnames
assert "Exceptions" in wb.sheetnames
print(f"  ✓ No Audit Log sheet (only {len(wb.sheetnames)} sheets)")

# Verify matched count matches API
matched_ws = wb["Matched Transactions"]
matched_rows = matched_ws.max_row - 1  # subtract header
api_matched = len(api_results["matched"])
print(f"\nMatched: Excel={matched_rows}, API={api_matched}")
assert matched_rows == api_matched, f"Mismatch! Excel={matched_rows} vs API={api_matched}"
print("  ✓ Matched count matches")

# Verify exceptions count matches API
exc_ws = wb["Exceptions"]
exc_rows = exc_ws.max_row - 1
api_exc = len(api_results["exceptions"])
print(f"Exceptions: Excel={exc_rows}, API={api_exc}")
assert exc_rows == api_exc, f"Mismatch! Excel={exc_rows} vs API={api_exc}"
print("  ✓ Exception count matches")

print("\nALL REPORT CHECKS PASSED!")
